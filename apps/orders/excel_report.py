"""
Generación de libro Excel (reporte de pedidos para administración).

Hojas:
- «Pedidos»: una fila por orden con cliente y totales.
- «Líneas»: una fila por ítem (toma, fechas, subtotal) con referencia al pedido.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _local_dt(dt) -> str:
    if dt is None:
        return ""
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime("%Y-%m-%d %H:%M")


def _money(value) -> float | str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return ""


def _autosize_columns(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 10
        for row in range(1, min(ws.max_row, 200) + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            ln = len(str(cell.value))
            if ln > best:
                best = min(ln + 2, max_width)
        ws.column_dimensions[letter].width = best


def build_orders_report_workbook(orders) -> Workbook:
    """
    Construye un libro con los pedidos del queryset (ya filtrado y con prefetch adecuado).
    """
    wb = Workbook()
    ws_o = wb.active
    ws_o.title = "Pedidos"
    ws_l = wb.create_sheet("Líneas")

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="FFF4F4F5")

    order_headers = [
        "ID pedido",
        "Código / referencia",
        "Estado (código)",
        "Estado",
        "Total USD",
        "Alta (creado)",
        "Enviado",
        "Reserva hasta",
        "Método de pago",
        "ID cliente",
        "Empresa",
        "RIF",
        "Contacto",
        "Email",
        "Teléfono",
        "Dirección",
        "Ciudad",
        "Estado cliente",
        "Workspace (slug)",
    ]
    for c, h in enumerate(order_headers, start=1):
        cell = ws_o.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    line_headers = [
        "ID pedido",
        "Código pedido",
        "ID línea",
        "Código toma",
        "Título toma",
        "Centro comercial",
        "Ciudad CC",
        "Inicio contrato",
        "Fin contrato",
        "Precio mensual USD",
        "Subtotal línea USD",
    ]
    for c, h in enumerate(line_headers, start=1):
        cell = ws_l.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    row_o = 2
    row_l = 2

    # Evaluar queryset (orden ya -created_at desde la vista)
    for order in orders:
        cl = order.client
        ws_o.cell(row=row_o, column=1, value=order.pk)
        ws_o.cell(row=row_o, column=2, value=_cell_str(order.code))
        ws_o.cell(row=row_o, column=3, value=order.status)
        ws_o.cell(row=row_o, column=4, value=order.get_status_display())
        ws_o.cell(row=row_o, column=5, value=_money(order.total_amount))
        ws_o.cell(row=row_o, column=6, value=_local_dt(order.created_at))
        ws_o.cell(row=row_o, column=7, value=_local_dt(order.submitted_at))
        ws_o.cell(row=row_o, column=8, value=_local_dt(order.hold_expires_at))
        ws_o.cell(row=row_o, column=9, value=order.get_payment_method_display())
        ws_o.cell(row=row_o, column=10, value=cl.pk if cl else "")
        ws_o.cell(row=row_o, column=11, value=_cell_str(cl.company_name if cl else ""))
        ws_o.cell(row=row_o, column=12, value=_cell_str(cl.rif if cl else ""))
        ws_o.cell(row=row_o, column=13, value=_cell_str(cl.contact_name if cl else ""))
        ws_o.cell(row=row_o, column=14, value=_cell_str(cl.email if cl else ""))
        ws_o.cell(row=row_o, column=15, value=_cell_str(cl.phone if cl else ""))
        ws_o.cell(row=row_o, column=16, value=_cell_str(cl.address if cl else ""))
        ws_o.cell(row=row_o, column=17, value=_cell_str(cl.city if cl else ""))
        ws_o.cell(row=row_o, column=18, value=cl.get_status_display() if cl else "")
        slug = ""
        if cl and getattr(cl, "workspace", None):
            slug = _cell_str(cl.workspace.slug)
        ws_o.cell(row=row_o, column=19, value=slug)
        row_o += 1

        for it in order.items.all():
            ad = it.ad_space
            sc = getattr(ad, "shopping_center", None)
            ws_l.cell(row=row_l, column=1, value=order.pk)
            ws_l.cell(row=row_l, column=2, value=_cell_str(order.code))
            ws_l.cell(row=row_l, column=3, value=it.pk)
            ws_l.cell(row=row_l, column=4, value=_cell_str(ad.code))
            ws_l.cell(row=row_l, column=5, value=_cell_str(ad.title))
            ws_l.cell(row=row_l, column=6, value=_cell_str(sc.name if sc else ""))
            ws_l.cell(row=row_l, column=7, value=_cell_str(sc.city if sc else ""))
            ws_l.cell(row=row_l, column=8, value=it.start_date.isoformat() if it.start_date else "")
            ws_l.cell(row=row_l, column=9, value=it.end_date.isoformat() if it.end_date else "")
            ws_l.cell(row=row_l, column=10, value=_money(it.monthly_price))
            ws_l.cell(row=row_l, column=11, value=_money(it.subtotal))
            row_l += 1

    _autosize_columns(ws_o)
    _autosize_columns(ws_l)
    return wb


def orders_report_excel_bytes(orders) -> bytes:
    """Serializa el reporte a bytes (.xlsx)."""
    buf = BytesIO()
    wb = build_orders_report_workbook(orders)
    wb.save(buf)
    return buf.getvalue()
